import urllib.request
import re
import json
import datetime
import csv
import os
from openpyxl import Workbook
from ucb import main

class get_stock_stats():
    def __init__(self, symbol, start_date, end_date):
        self.__symbol = symbol
        self.__start_date = start_date
        self.__end_date = end_date

    def __request(self, stat):
        url = 'http://finance.yahoo.com/d/quotes.csv?s=%s&f=%s' % (self.__symbol, stat)
        return urllib.request.urlopen(url).read().decode('utf8').strip().strip('"')

    def get_all(self):
        """
        Get all available quote data for the given ticker symbol.

        Returns a dictionary.
        """
        values = self.__request('l1c1va2xj1b4j4dyekjm3m4rr5p5p6s7').split(',')
        data = {}
        data['price'] = values[0]
        data['change'] = values[1]
        data['volume'] = values[2]
        data['avg_daily_volume'] = values[3]
        data['stock_exchange'] = values[4]
        data['market_cap'] = values[5]
        data['book_value'] = values[6]
        data['ebitda'] = values[7]
        data['dividend_per_share'] = values[8]
        data['dividend_yield'] = values[9]
        data['earnings_per_share'] = values[10]
        data['52_week_high'] = values[11]
        data['52_week_low'] = values[12]
        data['50day_moving_avg'] = values[13]
        data['200day_moving_avg'] = values[14]
        data['price_earnings_ratio'] = values[15]
        data['price_earnings_growth_ratio'] = values[16]
        data['price_sales_ratio'] = values[17]
        data['price_book_ratio'] = values[18]
        data['short_ratio'] = values[19]

        return data


    def get_historical_prices(self):
        """
        Get historical prices for the given ticker ymbol.
        Date format is 'YYYYMMDD'

        Returns a dictionary {date: {open: , high: , low: , close: , volume: , adjusted: }}
        Where date is in YYYY-MM-DD
        """
        url = 'http://ichart.yahoo.com/table.csv?s=%s&' % self.__symbol + \
            'd=%s&' % str(int(self.__end_date[4:6]) - 1) + \
            'e=%s&' % str(int(self.__end_date[6:8])) + \
            'f=%s&' % str(int(self.__end_date[0:4])) + \
            'g=d&' + \
            'a=%s&' % str(int(self.__start_date[4:6]) - 1) + \
            'b=%s&' % str(int(self.__start_date[6:8])) + \
            'c=%s&' % str(int(self.__start_date[0:4])) + \
            'ignore=.csv'
        days = urllib.request.urlopen(url).readlines()
        data = [day.decode('utf8')[:-2].split(',') for day in days]

        data_dic = {}
        for data_block in data[1:]:
            date_data = {}
            date_data['open'] = data_block[1]
            date_data['high'] = data_block[2]
            date_data['low'] = data_block[3]
            date_data['close'] = data_block[4]
            date_data['volume'] = data_block[5]
            date_data['adjusted'] = data_block[6]
            data_dic[data_block[0]] = date_data

        return data_dic

def get_tickers():
    ticker_list = []
    url = 'http://en.wikipedia.org/w/api.php?format=json&action=query&titles=List_of_S%26P_500_companies&prop=revisions&rvprop=content'

    f = urllib.request.urlopen(url)
    data = json.loads(f.read().decode('utf8'))
    data = data['query']['pages']['2676045']['revisions'][0]
    for key in data:
        data = data[key]

    tickers = re.findall('{{NyseSymbol\|\w+}}', data)
    tickers2 = re.findall('{{NasdaqSymbol\|\w+}}', data)

    for ticker in tickers:
        ticker_list.append(ticker[13:-2])
    for ticker in tickers2:
        ticker_list.append(ticker[15:-2])

    return ticker_list

class get_date():
    def __init__(self):
        date = datetime.datetime.today()
        self.__year = date.year
        self.__month = date.month
        self.__day = date.day

    def get_curr_date(self):
        year = str(self.__year)
        month = str(self.__month)
        day = str(self.__day)

        if month in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
            month = '0' + month
        if day in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
            day = '0' + day
        return year+month+day

    def get_prev_date(self):
        year = str(self.__year-3)
        month = str(self.__month)
        day = str(self.__day-1)

        if month in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
            month = '0' + month
        if day in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']:
            day = '0' + day
        return year+month+day

ticker_list = []

def get_curr_info():
    global ticker_list
    date = get_date()
    current = date.get_curr_date()
    previous = date.get_prev_date()
    if not ticker_list:
        ticker_list = get_tickers()
    stock_list = {}

    for ticker in ticker_list:
        try:
            print('Grabbing CURRENT data for %s' % ticker)
            ticker = ticker.lower()
            ticker_info = get_stock_stats(ticker, previous, current)
            curr_info = ticker_info.get_all()
            stock_list[ticker] = curr_info
        except urllib.error.HTTPError:
            print('404 Error for ' + ticker)

    return stock_list

def get_hist_info():
    global ticker_list
    date = get_date()
    current = date.get_curr_date()
    previous = date.get_prev_date()
    if not ticker_list:
        ticker_list = get_tickers()
    stock_list = {}

    for ticker in ticker_list:
        try:
            print('Grabbing HISTORICAL data for %s' % ticker)
            ticker = ticker.lower()
            ticker_info = get_stock_stats(ticker, previous, current)
            hist_info = ticker_info.get_historical_prices()
            stock_list[ticker] = hist_info
        except urllib.error.HTTPError:
            print('404 Error for ' + ticker)

    return stock_list

def current_data():
    data = get_curr_info()
    f = Workbook()
    s = f.get_active_sheet()
    k = 0
    for ticker in data:
        print('Adding {0} (#{1}) to CURRENT spreadsheet'.format(ticker, k+1))
        s.title = ticker.upper()
        s.cell('A1').value = 'symbol'
        s.cell('A2').value = ticker.upper()
        i = 1
        for key in data[ticker]:
            s.cell(row = 0, column = i).value = key
            s.cell(row = 1, column = i).value = data[ticker][key]
            i += 1
        f.save('sp500_curr.xls')
        k += 1
        s = f.create_sheet(title = 'new_sheet')
        s = f.get_sheet_by_name('new_sheet')
    f.save('sp500_curr.xls')

def historical_data():
    data = get_hist_info()
    f = Workbook()
    s = f.get_active_sheet()
    k = 0
    for ticker in data:
        print('Adding {0} (#{1}) to HISTORICAL spreadsheet'.format(ticker, k+1))
        s.title = ticker.upper()
        s.cell('A1').value = 'symbol'
        s.cell('A2').value = ticker.upper()
        s.cell('B1').value = 'date'
        s.cell('C1').value = 'volume'
        s.cell('D1').value = 'adjusted'
        s.cell('E1').value = 'high'
        s.cell('F1').value = 'low'
        s.cell('G1').value = 'close'
        s.cell('F1').value = 'open'
        i = 1
        j = 1
        for date in data[ticker]:
            s.cell(row = j, column = i).value = date
            for key in data[ticker][date]:
                i += 1
                s.cell(row = j, column = i).value = data[ticker][date][key]
            i -= 6
            j += 1
        f.save('sp500_hist.xls')
        k += 1
        s = f.create_sheet(title = 'new_sheet')
        s = f.get_sheet_by_name('new_sheet')
    f.save('sp500_hist.xls')












